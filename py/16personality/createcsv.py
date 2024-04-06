# excelファイル内の各シートをcsvファイルに変換する

import pandas as pd
from openpyxl import load_workbook

# excel ファイルを開いて、シート情報を取得する
wb = load_workbook('202404-appteam16Personality.xlsx')
wb_sheet_names = wb.sheetnames
print(wb_sheet_names)

sheet_source = wb['source']
sheet_master_category = wb['master-category']

# category マスタ辞書作成
master_category = {}
for row in sheet_master_category.iter_rows(min_row=2, max_row=sheet_master_category.max_row, min_col=1, max_col=sheet_master_category.max_column):
    master_category[row[1].value] = row[0].value

print(master_category)

# テーブルデータの作成
personalities = []
for row in sheet_source.iter_rows(min_row=2, max_row=sheet_source.max_row, min_col=1, max_col=sheet_source.max_column):
    # カテゴリ名はカテゴリマスタシートから対応するものを取得する
    category_name = master_category[row[1].value]

    # エネルギー・意識・気質・戦術・identity それぞれのデータを作成する。ペアになる要素は二つ合計して100になるようにする
    # レート情報がない場合は要素値を0にする
    energy = row[2].value
    energy_rate = row[3].value
    e = 0
    i = 0
    if energy_rate != None:
        energy_rate = int(energy_rate)

        if energy == 'E':
            e = energy_rate
            i = 100 - e
        elif energy == 'I':
            i = energy_rate
            e = 100 - i
        else:
            print('error-energy')
            exit()
    
    consciousness = row[4].value
    consciousness_rate = row[5].value
    n = 0
    s = 0
    if consciousness_rate != None:
        consciousness_rate = int(consciousness_rate)

        if consciousness == 'N':
            n = consciousness_rate
            s = 100 - n
        elif consciousness == 'S':
            s = consciousness_rate
            n = 100 - s
        else:
            print('error-consciousness')
            exit()

    charactor = row[6].value
    charactor_rate = row[7].value
    t = 0
    f = 0
    if charactor_rate != None:
        charactor_rate = int(charactor_rate)
    
        if charactor == 'T':
            t = charactor_rate
            f = 100 - t
        elif charactor == 'F':
            f = charactor_rate
            t = 100 - f
        else:
            print('error-charactor')
            exit()

    tactic = row[8].value
    tactic_rate = row[9].value
    j = 0
    p = 0
    if tactic_rate != None:
        tactic_rate = int(tactic_rate)
    
        if tactic == 'J':
            j = tactic_rate
            p = 100 - j
        elif tactic == 'P':
            p = tactic_rate
            j = 100 - p
        else:
            print('error-tactic')
            exit()

    identity = row[10].value
    identity_rate = row[11].value
    cat_a = 0
    cat_t = 0
    if identity_rate != None:
        identity_rate = int(identity_rate)
    
        if identity == 'A':
            cat_a = identity_rate
            cat_t = 100 - cat_a
        elif identity == 'T':
            cat_t = identity_rate
            cat_a = 100 - cat_t
        else:
            print('error-identity')
            print(identity)
    
    personalities.append([row[0].value, category_name, identity, e, i, n, s, t, f, j, p, cat_a, cat_t])

# personality シートにデータを書き込む
if 'personality' in wb.sheetnames:
    wb.remove(wb['personality'])

sheet_personality = wb.create_sheet('personality')
header_personality = ['name', 'category', 'identity', 'E', 'I', 'N', 'S', 'T', 'F', 'J', 'P', 'cat-A', 'cat-T']

cnt = 1

sheet_personality['A' + str(cnt)] = header_personality[0]
sheet_personality['B' + str(cnt)] = header_personality[1]
sheet_personality['C' + str(cnt)] = header_personality[2]
sheet_personality['D' + str(cnt)] = header_personality[3]
sheet_personality['E' + str(cnt)] = header_personality[4]
sheet_personality['F' + str(cnt)] = header_personality[5]
sheet_personality['G' + str(cnt)] = header_personality[6]
sheet_personality['H' + str(cnt)] = header_personality[7]
sheet_personality['I' + str(cnt)] = header_personality[8]
sheet_personality['J' + str(cnt)] = header_personality[9]
sheet_personality['K' + str(cnt)] = header_personality[10]
sheet_personality['L' + str(cnt)] = header_personality[11]
sheet_personality['M' + str(cnt)] = header_personality[12]

cnt += 1

for personality in personalities:
    sheet_personality['A' + str(cnt)] = personality[0]
    sheet_personality['B' + str(cnt)] = personality[1]
    sheet_personality['C' + str(cnt)] = personality[2]
    sheet_personality['D' + str(cnt)] = personality[3]
    sheet_personality['E' + str(cnt)] = personality[4]
    sheet_personality['F' + str(cnt)] = personality[5]
    sheet_personality['G' + str(cnt)] = personality[6]
    sheet_personality['H' + str(cnt)] = personality[7]
    sheet_personality['I' + str(cnt)] = personality[8]
    sheet_personality['J' + str(cnt)] = personality[9]
    sheet_personality['K' + str(cnt)] = personality[10]
    sheet_personality['L' + str(cnt)] = personality[11]
    sheet_personality['M' + str(cnt)] = personality[12]

    cnt += 1

wb.save('202404-appteam16Personality.xlsx')
wb.close()

# csv ファイルに変換する
df_personality = pd.read_excel('202404-appteam16Personality.xlsx', sheet_name='personality')
df_personality.to_csv('personality.csv', index=False)

df_master_category = pd.read_excel('202404-appteam16Personality.xlsx', sheet_name='master-category')
df_master_category.to_csv('master-category.csv', index=False)

df_master_identity = pd.read_excel('202404-appteam16Personality.xlsx', sheet_name='master-identity')
df_master_identity.to_csv('master-identity.csv', index=False)

df_master_element = pd.read_excel('202404-appteam16Personality.xlsx', sheet_name='master-element')
df_master_element.to_csv('master-element.csv', index=False)

print('finish')